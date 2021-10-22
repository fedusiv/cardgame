from openpyxl import load_workbook

from cards.card_enums import CardObjType
from cards.card_obj import CardObj

class CardPoolBuilder:

    def __init__(self):
        pass

    def build_from_excel(self, filepath):
        card_list = []
        wb = load_workbook(filename=filepath)
        sheet_deck = wb["Deck"]
        rows_amount = sheet_deck.max_row - 1  # -1 because one row is used for description
        # Call the builder
        for r in range(self.row_start_index, self.row_start_index + rows_amount):
            card = self.parse_row(r, sheet_deck)
            if card is not None:
                card_list.append(card)
        return card_list

    # Parsing part
    row_start_index = 2
    column_id = 'A'
    column_card_type = 'B'
    column_name = 'C'
    column_cost_action = 'D'
    column_cost_mana = 'E'

    def parse_row(self, row_id, sheet_deck):
        # Card type parsing
        card_type: CardObjType
        card_type_str = sheet_deck[self.column_card_type + str(row_id)].value
        if card_type_str == "Mana":
            card_type = CardObjType.Mana
        elif card_type_str == "Spell":
            card_type = CardObjType.Spell
        else:
            # Type of object is unspecified, do not parse it
            return None
        # Unique id for all cards
        # TODO: add verification, that this id is really unique
        card_id = sheet_deck[self.column_id + str(row_id)].value
        # Name of card
        name = sheet_deck[self.column_name + str(row_id)].value
        # Price
        cost_action = sheet_deck[self.column_cost_action + str(row_id)].value
        cost_mana = sheet_deck[self.column_cost_mana + str(row_id)].value
        card = CardObj(card_id, name, cost_action, cost_mana, card_type)
        return card

