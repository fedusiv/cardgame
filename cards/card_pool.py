from openpyxl import load_workbook

class CardPool:
    # Singleton part
    _instance = None

    @staticmethod
    def instance():
        if CardPool._instance is None:
            CardPool()
        return CardPool._instance

    def __init__(self):
        CardPool.__instance = self
        self.parse_db()

    # Parsing part
    row_start_index = 2
    column_id = 'A'
    column_card_type = 'B'
    column_name = 'C'
    column_price = 'D'
    def parse_db(self):
        wb = load_workbook(filename='cards_description/graveyard.xlsx')
        sheet_deck = wb["Deck"]
        rows_amount = sheet_deck.max_row - 1    # -1 because one row is used for description
        for r in range(self.row_start_index, self.row_start_index+rows_amount):
            self.parse_row(r, sheet_deck)

    def parse_row(self, row_id, sheet_deck):
        card_id = sheet_deck[self.column_id+str(row_id)].value
        name = sheet_deck[self.column_name+str(row_id)].value


